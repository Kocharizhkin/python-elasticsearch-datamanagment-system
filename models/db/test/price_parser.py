# This Python file uses the following encoding: utf-8
# parse price-lists and inserts them in uniform format into a single file
# insert_price_into_global_price(file_to_read, file_to_write, code_col_num, isbn_col_num, author_col_num, title_col_num, publ_col_num, price_col_num, starting_row, discount, to_sheet_name) :

import openpyxl
from openpyxl import Workbook
from datetime import datetime
import ftplib
import zipfile
import rarfile
import os
import re
import pathlib
from pathlib import Path
from time import sleep
#from typing import NamedTuple


	

def isbn2shortean(isbn) : #transforming isbn-13 and isbn-10 to isbn-9 (first 9 meaningful characters)
	if len(str(isbn)) < 9 :
		ean = '000' # arbitrary equivalent marking invalid isbns (issns, cut, missing etc)
	else :
		ean = re.sub('[^0-9]', '', isbn)
		ean = ean[-10:-1]
	return str(ean)

def insert_price_into_global_price(file_to_read, file_to_write, code_col_num, isbn_col_num, author_col_num, title_col_num, publ_col_num, price_col_num, starting_row, discount, to_sheet_name) :

	wb = openpyxl.load_workbook(file_to_read, read_only=True, data_only=True)	
	ws = wb.active
	maxrow = ws.max_row
	data = []
	row_count = 1

	print('reading ',maxrow, ' lines of ',file_to_read)
	for row in ws.iter_rows(min_row=starting_row, max_row=maxrow) :
		if row[code_col_num-1].value != None :
			data.append(['' if cell.value == None else cell.value for cell in row])
		print('\r', round(round(row_count/maxrow,2)*100,0),'%',end='')
		row_count = row_count + 1

	wb.close()
	print()

	if Path(working_dir+'\\'+file_to_write).exists() and Path(working_dir+'\\'+file_to_write).is_file() :
		print(Path(working_dir+'\\'+file_to_write), 'exists, checking a worksheet ', to_sheet_name, ' there')
		wb1 = openpyxl.load_workbook(file_to_write, data_only=True)
		if not to_sheet_name in wb1.sheetnames :
			print('adding a worksheet ', to_sheet_name)
			ws = wb1.create_sheet(to_sheet_name)
		else :
			print('clearing the existing worksheet ', to_sheet_name)
			wb1.remove(wb1[to_sheet_name])
			ws = wb1.create_sheet(to_sheet_name)				
	else :
		print('creating a new file ', file_to_write)
		wb1 = Workbook(write_only=True)
		ws = wb1.create_sheet(to_sheet_name)

	good_row = []
	print('writing ', len(data), ' lines to ',file_to_write)
	for i in range(1, len(data)) :
		try :
			float(data[i-1][price_col_num-1])
			good_row = [isbn2shortean(str(data[i-1][isbn_col_num-1])), str(data[i-1][code_col_num-1])+' | '+str(data[i-1][isbn_col_num-1])+' | ' + str(data[i-1][author_col_num-1])+' '+ str(data[i-1][title_col_num-1])+' '+str(data[i-1][publ_col_num-1])+' #'+str(round((100-discount)/100*round(float(data[i-1][price_col_num-1]),2),2)).replace('.',','), str(data[i-1][code_col_num-1]), str(data[i-1][isbn_col_num-1]), str(data[i-1][author_col_num-1]), str(data[i-1][title_col_num-1]), str(data[i-1][publ_col_num-1]), str(round((100-discount)/100*round(float(data[i-1][price_col_num-1]),2),2)).replace('.',',')]
		except :
			good_row = [isbn2shortean(str(data[i-1][isbn_col_num-1])), str(data[i-1][code_col_num-1])+' | '+str(data[i-1][isbn_col_num-1])+' | ' + str(data[i-1][author_col_num-1])+' '+ str(data[i-1][title_col_num-1])+' '+str(data[i-1][publ_col_num-1])+' #99999', str(data[i-1][code_col_num-1]), str(data[i-1][isbn_col_num-1]), str(data[i-1][author_col_num-1]), str(data[i-1][title_col_num-1]), str(data[i-1][publ_col_num-1]), '99999']
	
#		good_row[:] = ['' if cell_x == None else cell_x for cell_x in good_row]

		ws.append(good_row)
		print('\r', round(round(i/len(data),2)*100,0),'%',end='')


	print()		
	wb1.save(file_to_write)
	wb1.close()
	return None














working_dir = r'D:\files\mk\filtering\tmp'

isbn0 = ''
file_to = r'price_filtering_prices.xlsx'

os.chdir(working_dir)
print('getting to working directory: ', os.getcwd())
#cook_raw_prices(Path(working_dir+r'\tmp'))

#file_to_read, file_to_write, code_col_num, isbn_col_num, author_col_num, title_col_num, publ_col_num, price_col_num, starting_row, discount, to_sheet_name

insert_price_into_global_price('hit_price_ast.xlsx', file_to, 1, 10, 4, 5, 13, 14, 11, 20, 'hit1')
insert_price_into_global_price('hit_price_assort.xlsx', file_to, 1, 10, 4, 5, 13, 14, 11, 20, 'hit2')
insert_price_into_global_price('hit_price_uch.xlsx', file_to, 1, 10, 4, 5, 13, 14, 11, 20, 'hit3')
insert_price_into_global_price('366_price.xlsx', file_to, 2, 8, 3, 4, 5, 7, 9, 18, '366')
insert_price_into_global_price('omega_price.xlsx', file_to, 1, 9, 3, 2, 4, 7, 2, 0, 'omega')
#insert_price_into_global_price('omega_full_price.xlsx', file_to, 1, 9, 3, 2, 4, 7, 2, 0, 'omega_full')
#insert_price_into_global_price('omega_discount_price.xlsx', file_to, 1, 21, 7, 2, 8, 6, 3, 0, 'omega_discount')
#insert_price_into_global_price('infra_price.xlsx', file_to, 2, 13, 6, 4, 8, 3, 8, 0, 'infra')
#insert_price_into_global_price('soyuz_price.xlsx', file_to, 15, 16, 4, 3, 6, 2, 8, 10, 'soyuz')
#insert_price_into_global_price('gnosis_shop_price.xlsx', file_to, 1, 7, 2, 3, 11, 13, 12, 0, 'gnosis_shop')
insert_price_into_global_price('gnosis_price.xlsx', file_to, 2, 10, 3, 4, 15, 6, 11, 0, 'gnosis')
insert_price_into_global_price('gnosis_price.xlsx', file_to, 2, 10, 3, 4, 15, 6, 11, 0, 'gnosis1')
#insert_price_into_global_price('kraft_price.xlsx', file_to, 1, 10, 2, 3, 5, 11, 2, 0, 'kraft')
#insert_price_into_global_price('knart_price.xlsx', file_to, 1, 2, 3, 4, 5, 11, 2, 0, 'knart')
#insert_price_into_global_price('uk_price.xlsx', file_to, 1, 7, 10, 2, 3, 5, 2, 0, 'uk')
#insert_price_into_global_price('mif_price.xlsx', file_to, 1, 7, 6, 5, 15, 3, 2, 40, 'mif')
#insert_price_into_global_price('gumak_price.xlsx', file_to, 1, 16, 3, 4, 5, 15, 2, 0, 'gumak')